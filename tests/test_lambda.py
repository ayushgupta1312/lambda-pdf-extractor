"""
Unit tests for the Lambda PDF extractor function.
"""

import json
import os
import sys
from io import BytesIO
from unittest.mock import MagicMock, patch

import pytest

# Add src to path for imports
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

from lambda_function import (
    create_excel_from_tables,
    create_response,
    extract_tables_from_pdf,
    lambda_handler,
)


class TestCreateResponse:
    """Tests for the create_response function."""

    def test_create_response_success(self):
        """Test successful response creation."""
        response = create_response(200, "Success")
        assert response["statusCode"] == 200
        body = json.loads(response["body"])
        assert body["message"] == "Success"

    def test_create_response_error(self):
        """Test error response creation."""
        response = create_response(500, "Error occurred")
        assert response["statusCode"] == 500
        body = json.loads(response["body"])
        assert body["message"] == "Error occurred"


class TestExtractTablesFromPdf:
    """Tests for the extract_tables_from_pdf function."""

    def test_extract_tables_from_sample_pdf(self):
        """Test table extraction from the sample PDF."""
        sample_pdf_path = os.path.join(
            os.path.dirname(__file__), "..", "sample-input", "sample_data.pdf"
        )

        if os.path.exists(sample_pdf_path):
            with open(sample_pdf_path, "rb") as f:
                pdf_content = f.read()

            tables = extract_tables_from_pdf(pdf_content)

            # Should extract at least one table
            assert len(tables) >= 1

            # Tables should not be empty
            for table in tables:
                assert len(table) > 0
                for row in table:
                    assert isinstance(row, list)

    def test_extract_tables_handles_empty_cells(self):
        """Test that None values are converted to empty strings."""
        sample_pdf_path = os.path.join(
            os.path.dirname(__file__), "..", "sample-input", "sample_data.pdf"
        )

        if os.path.exists(sample_pdf_path):
            with open(sample_pdf_path, "rb") as f:
                pdf_content = f.read()

            tables = extract_tables_from_pdf(pdf_content)

            # All cells should be strings
            for table in tables:
                for row in table:
                    for cell in row:
                        assert isinstance(cell, str)


class TestCreateExcelFromTables:
    """Tests for the create_excel_from_tables function."""

    def test_create_excel_single_table(self):
        """Test Excel creation with a single table."""
        tables = [
            [["Header1", "Header2"], ["Value1", "Value2"], ["Value3", "Value4"]]
        ]

        excel_content = create_excel_from_tables(tables)

        # Should return bytes
        assert isinstance(excel_content, bytes)

        # Should be a valid Excel file (starts with PK for ZIP/XLSX format)
        assert excel_content[:2] == b"PK"

    def test_create_excel_multiple_tables(self):
        """Test Excel creation with multiple tables."""
        tables = [
            [["A1", "B1"], ["A2", "B2"]],
            [["X1", "Y1", "Z1"], ["X2", "Y2", "Z2"]],
        ]

        excel_content = create_excel_from_tables(tables)

        assert isinstance(excel_content, bytes)
        assert excel_content[:2] == b"PK"

    def test_create_excel_empty_table(self):
        """Test Excel creation with empty cells."""
        tables = [[["", "Header"], ["Value", ""]]]

        excel_content = create_excel_from_tables(tables)

        assert isinstance(excel_content, bytes)
        assert excel_content[:2] == b"PK"


class TestLambdaHandler:
    """Tests for the main lambda_handler function."""

    def test_handler_no_records(self):
        """Test handler with no records in event."""
        event = {"Records": []}
        response = lambda_handler(event, None)

        assert response["statusCode"] == 400

    def test_handler_empty_event(self):
        """Test handler with empty event."""
        event = {}
        response = lambda_handler(event, None)

        assert response["statusCode"] == 400

    def test_handler_non_pdf_file(self):
        """Test handler skips non-PDF files."""
        event = {
            "Records": [
                {
                    "s3": {
                        "bucket": {"name": "magnifact-pdf"},
                        "object": {"key": "input-pdf-files/document.txt"},
                    }
                }
            ]
        }

        with patch("lambda_function.s3_client") as mock_s3:
            response = lambda_handler(event, None)

        # Should succeed but skip the file
        assert response["statusCode"] == 200
        # S3 should not be called for non-PDF files
        mock_s3.get_object.assert_not_called()

    def test_handler_wrong_folder(self):
        """Test handler skips files in wrong folder."""
        event = {
            "Records": [
                {
                    "s3": {
                        "bucket": {"name": "magnifact-pdf"},
                        "object": {"key": "other-folder/document.pdf"},
                    }
                }
            ]
        }

        with patch("lambda_function.s3_client") as mock_s3:
            response = lambda_handler(event, None)

        # Should succeed but skip the file
        assert response["statusCode"] == 200
        mock_s3.get_object.assert_not_called()

    @patch("lambda_function.s3_client")
    def test_handler_successful_processing(self, mock_s3):
        """Test successful PDF processing."""
        sample_pdf_path = os.path.join(
            os.path.dirname(__file__), "..", "sample-input", "sample_data.pdf"
        )

        if not os.path.exists(sample_pdf_path):
            pytest.skip("Sample PDF not found")

        with open(sample_pdf_path, "rb") as f:
            pdf_content = f.read()

        # Mock S3 get_object response
        mock_body = MagicMock()
        mock_body.read.return_value = pdf_content
        mock_s3.get_object.return_value = {"Body": mock_body}

        event = {
            "Records": [
                {
                    "s3": {
                        "bucket": {"name": "magnifact-pdf"},
                        "object": {"key": "input-pdf-files/sample_data.pdf"},
                    }
                }
            ]
        }

        response = lambda_handler(event, None)

        assert response["statusCode"] == 200

        # Verify S3 was called correctly
        mock_s3.get_object.assert_called_once_with(
            Bucket="magnifact-pdf", Key="input-pdf-files/sample_data.pdf"
        )

        # Verify put_object was called with correct output path
        put_call = mock_s3.put_object.call_args
        assert put_call[1]["Bucket"] == "magnifact-pdf"
        assert put_call[1]["Key"] == "output-files/sample_data.xlsx"
        assert (
            put_call[1]["ContentType"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


class TestIntegration:
    """Integration tests using the sample PDF."""

    def test_full_extraction_pipeline(self):
        """Test the complete extraction pipeline."""
        sample_pdf_path = os.path.join(
            os.path.dirname(__file__), "..", "sample-input", "sample_data.pdf"
        )

        if not os.path.exists(sample_pdf_path):
            pytest.skip("Sample PDF not found")

        # Read sample PDF
        with open(sample_pdf_path, "rb") as f:
            pdf_content = f.read()

        # Extract tables
        tables = extract_tables_from_pdf(pdf_content)
        assert len(tables) >= 1

        # Create Excel
        excel_content = create_excel_from_tables(tables)
        assert len(excel_content) > 0

        # Verify it's a valid XLSX file
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(excel_content))
        assert len(wb.sheetnames) >= 1

        # Verify sheets have data
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            assert sheet.max_row > 0
            assert sheet.max_column > 0
