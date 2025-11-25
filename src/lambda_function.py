"""
AWS Lambda function to extract tables from PDF files and convert them to Excel format.

This Lambda function is triggered by S3 events when a new PDF file is uploaded to
the configured input folder. It extracts tables from the PDF using pdfplumber and
saves them to an Excel file in the output folder.

Environment Variables:
    BUCKET_NAME: S3 bucket name (default: magnifact-pdf)
    INPUT_FOLDER_NAME: Input folder for PDF files (default: input-pdf-files)
    OUTPUT_FOLDER_NAME: Output folder for Excel files (default: output-files)
"""

import json
import logging
import os
import urllib.parse
from io import BytesIO
from typing import Any

import boto3
import pdfplumber
from botocore.exceptions import ClientError
from openpyxl import Workbook

# Configure logging
logger = logging.getLogger()
logger.setLevel(logging.INFO)

# Environment variables with defaults
BUCKET_NAME = os.environ.get("BUCKET_NAME", "magnifact-pdf")
INPUT_FOLDER_NAME = os.environ.get("INPUT_FOLDER_NAME", "input-pdf-files")
OUTPUT_FOLDER_NAME = os.environ.get("OUTPUT_FOLDER_NAME", "output-files")

# Initialize S3 client
s3_client = boto3.client("s3")


def lambda_handler(event: dict[str, Any], context: Any) -> dict[str, Any]:
    """
    Main Lambda handler function.

    Args:
        event: S3 event notification containing information about the uploaded file
        context: Lambda context object

    Returns:
        Response dictionary with status code and message
    """
    logger.info("Received event: %s", json.dumps(event, indent=2))

    try:
        # Extract S3 bucket and key from the event
        records = event.get("Records", [])
        if not records:
            logger.warning("No records found in the event")
            return create_response(400, "No records found in event")

        for record in records:
            # Get bucket name and object key from the event
            bucket = record["s3"]["bucket"]["name"]
            key = urllib.parse.unquote_plus(record["s3"]["object"]["key"])

            logger.info("Processing file: %s from bucket: %s", key, bucket)

            # Validate that the file is in the input folder
            if not key.startswith(f"{INPUT_FOLDER_NAME}/"):
                logger.info("File %s is not in the input folder, skipping", key)
                continue

            # Validate that the file is a PDF
            if not key.lower().endswith(".pdf"):
                logger.info("File %s is not a PDF file, skipping", key)
                continue

            # Process the PDF file
            process_pdf(bucket, key)

        return create_response(200, "PDF processing completed successfully")

    except ClientError as e:
        logger.error("AWS Client Error: %s", str(e))
        return create_response(500, f"AWS Error: {str(e)}")
    except Exception as e:
        logger.error("Error processing PDF: %s", str(e))
        return create_response(500, f"Error: {str(e)}")


def process_pdf(bucket: str, key: str) -> None:
    """
    Download PDF from S3, extract tables, and upload Excel file.

    Args:
        bucket: S3 bucket name
        key: S3 object key (path to the PDF file)
    """
    # Download PDF from S3
    logger.info("Downloading PDF from S3: %s/%s", bucket, key)
    pdf_object = s3_client.get_object(Bucket=bucket, Key=key)
    pdf_content = pdf_object["Body"].read()

    # Extract tables from PDF
    logger.info("Extracting tables from PDF")
    tables = extract_tables_from_pdf(pdf_content)

    if not tables:
        logger.warning("No tables found in PDF: %s", key)
        # Create an empty Excel file with a message
        tables = [[["No tables found in the PDF file"]]]

    # Create Excel workbook
    logger.info("Creating Excel workbook")
    excel_content = create_excel_from_tables(tables)

    # Generate output key
    filename = os.path.basename(key)
    filename_without_ext = os.path.splitext(filename)[0]
    output_key = f"{OUTPUT_FOLDER_NAME}/{filename_without_ext}.xlsx"

    # Upload Excel to S3
    logger.info("Uploading Excel to S3: %s/%s", bucket, output_key)
    s3_client.put_object(
        Bucket=bucket,
        Key=output_key,
        Body=excel_content,
        ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    logger.info("Successfully processed %s and uploaded %s", key, output_key)


def extract_tables_from_pdf(pdf_content: bytes) -> list[list[list[str]]]:
    """
    Extract all tables from a PDF file.

    Args:
        pdf_content: PDF file content as bytes

    Returns:
        List of tables, where each table is a list of rows,
        and each row is a list of cell values
    """
    tables = []
    pdf_stream = BytesIO(pdf_content)

    with pdfplumber.open(pdf_stream) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            logger.info("Processing page %d of %d", page_num, len(pdf.pages))

            # Extract tables from the page
            page_tables = page.extract_tables()

            for table in page_tables:
                if table:
                    # Clean table data - replace None values with empty strings
                    cleaned_table = [
                        [str(cell) if cell is not None else "" for cell in row]
                        for row in table
                    ]
                    tables.append(cleaned_table)

    logger.info("Extracted %d tables from PDF", len(tables))
    return tables


def create_excel_from_tables(tables: list[list[list[str]]]) -> bytes:
    """
    Create an Excel workbook from extracted tables.

    Args:
        tables: List of tables extracted from PDF

    Returns:
        Excel file content as bytes
    """
    workbook = Workbook()

    # Remove default sheet if we have tables to add
    if tables:
        default_sheet = workbook.active
        workbook.remove(default_sheet)

    for idx, table in enumerate(tables, start=1):
        # Create a new sheet for each table
        sheet_name = f"Table_{idx}"
        worksheet = workbook.create_sheet(title=sheet_name)

        for row_idx, row in enumerate(table, start=1):
            for col_idx, cell_value in enumerate(row, start=1):
                worksheet.cell(row=row_idx, column=col_idx, value=cell_value)

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width

    # Save workbook to bytes
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output.getvalue()


def create_response(status_code: int, message: str) -> dict[str, Any]:
    """
    Create a standardized response dictionary.

    Args:
        status_code: HTTP status code
        message: Response message

    Returns:
        Response dictionary
    """
    return {
        "statusCode": status_code,
        "body": json.dumps({"message": message}),
    }
